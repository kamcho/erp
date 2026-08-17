"""
Import Excel Grassland students from Term 2 2026 fee workbook.

Expected workbook layout (one sheet per class stream):
  - Column A: admission number
  - Column C (NAME/NAMES header): student name
  - Last balance column (header BAL when present): opening fee balance

Usage:
  python manage.py import_excel_grassland_students --file "c:\\term 2 2026.xlsx"
  python manage.py import_excel_grassland_students --file /path/to/file.xlsx --clear
"""
from __future__ import annotations

import re
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction

from core.models import Class, Grade, School, Student, StudentProfile

SCHOOL_NAME = "Excel Grassland"
DEFAULT_FILE = Path(r"c:\term 2 2026.xlsx")

SHEET_CLASS_MAP = {
    "pg": ("Play Group", "Play Group"),
    "pp1": ("PP1", "PP1"),
    "pp2": ("PP2", "PP2"),
    "gr1": ("Grade 1", "East"),
    "gr2": ("Grade 2", "Grade 2"),
    "3w": ("Grade 3", "West"),
    "3e": ("Grade 3", "East"),
    "4w": ("Grade 4", "Grade 4"),
    "4e": ("Grade 4", "Grade 4"),
    "5": ("Grade 5", "Grade 5"),
    "6": ("Grade 6", "Grade 6"),
    "7a": ("Grade 7", "Amazon"),
    "7e": ("Grade 7", "Everest"),
    "8": ("Grade 8", "Grade 8"),
    "9": ("Grade 9", "Everest"),
}

DOB_BY_GRADE = {
    "Play Group": date(2022, 6, 1),
    "PP1": date(2021, 6, 1),
    "PP2": date(2020, 6, 1),
    "Grade 1": date(2019, 6, 1),
    "Grade 2": date(2018, 6, 1),
    "Grade 3": date(2017, 6, 1),
    "Grade 4": date(2016, 6, 1),
    "Grade 5": date(2015, 6, 1),
    "Grade 6": date(2014, 6, 1),
    "Grade 7": date(2013, 6, 1),
    "Grade 8": date(2012, 6, 1),
    "Grade 9": date(2011, 6, 1),
}
JOINED = date(2026, 1, 5)

FEMALE_FIRST = {
    "TIFFANY", "NEEMA", "HEAVENLY", "GRACIOUS", "LUCIL", "LEILANI", "ABIGAIL",
    "ALMA", "AYLA", "SUCCESS", "IMANI", "MILKA", "ZELENE", "ESTHER", "SHIRLEEN",
    "YASMIN", "AVI", "JOAN", "SHANTEL", "HADASSA", "VALENCIA", "JANET", "BEBORA",
    "MARY", "JOANNE", "IVY", "GRACE", "MITCHELL", "CELLITA", "VARINE", "VICTORIA",
    "MARGARET", "TASHA", "ALICIA", "LATTIFA", "FAITH", "AMARA", "PRECIOUS",
    "ARIELLA", "BELLA", "NEPHRINE", "GIFT", "JOY", "NAOMI", "ANN", "LILIAN",
    "JOYLYN", "JOYLIZIA", "JECINTA", "ASHLEY", "SUSAN", "DEANE", "MARY", "MILLICENT",
    "TRIVIA", "ESSY", "LIBERTY", "MARGARET", "NATASHA", "BRINLEY", "TRIZA", "NELVINE",
    "MERCY", "WENDY", "DAMIELLA", "ZAYN", "ZUENA", "FRANCICA", "TINSLEE", "ANGELLIZ",
    "MILAN", "PEACE", "FAVOUR", "EVE", "KELLY", "ASHLYN", "ZAIRA", "TALIA", "DAISY",
    "BLESSING", "IMMACULATE", "JASMINE", "WANGUI", "MELISSA", "PENELOPE", "CIENNA",
    "PRINCESS", "BERNICE", "PERPETUA", "MONICA", "KYLAH", "SARAH", "SHAYNE", "ANGEL",
    "MALIA", "AMIRA", "CINDY", "JENNIFER", "ZAINAB", "IMELA", "OLIVE", "DESTINY",
    "MARISSA", "VANESSA", "LISAH", "STACY", "TAMARA", "TATIANA", "BIANNA", "GABRIELLA",
    "MEGAN", "PAULINE", "ATTIA", "TRIZAH", "SUBIRA", "BRIDGET", "RAHMA", "KAITLINE",
    "LOVELYNE", "AMAYA", "SHALOM", "MUNIRA", "SASHA", "TERESA", "UMULKHEIR", "LYNN",
    "IRENE", "SHERRY", "NAJRA", "TRACY", "VELMA", "SHANNEL", "DEONNE", "SAMIRA",
    "TABITHA", "JAMELIA", "CAITLYNN", "CANDICE", "MELISAH", "BRITNEY", "RENEE",
    "STEPHANIE", "ARIEL", "REGINA", "SPECIOSA", "SAMARA", "RACHAEL", "VALARIE", "JANE",
    "LUCY", "NATASHA", "JOYLINE", "ALEXIS", "NERIA", "VICKY", "HANIFA", "SHARLEEN",
    "PENDO", "LISA", "MEGHAN", "NADIA", "ALISHA", "DAVINA", "CELINE", "RUBY",
    "SHIRLEY", "ASHLY", "ANSELEM", "AMADER", "REIGNA", "WANGARI", "WANJIRU", "WANJIKU",
    "WAIRIMU", "WAMBUI", "NJOKI", "NYOKABI", "CHEPKOECH", "CHELANGAT", "AKOTH",
    "WANJIRU", "ATIENO", "WANGECI", "KENDI", "CHEMUTAI", "MUTHONI", "WACERA",
    "WANJIKU", "NJERI", "NYAMBURA", "WAITHIRA",
}


def split_name(full: str) -> tuple[str, str, str]:
    parts = [p for p in re.sub(r"\s+", " ", full.strip()).split() if p]
    if not parts:
        return "Unknown", "", "Unknown"
    if len(parts) == 1:
        return parts[0].title(), "", parts[0].title()
    if len(parts) == 2:
        return parts[0].title(), "", parts[1].title()
    return parts[0].title(), " ".join(parts[1:-1]).title(), parts[-1].title()


def guess_gender(first: str) -> str:
    return "female" if first.upper() in FEMALE_FIRST else "male"


def parse_balance(value) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(round(float(value)))
    text = str(value).strip().upper()
    if text in {"NIL", "NILL", "-", "—", "NONE"}:
        return 0
    text = text.replace(",", "")
    try:
        return int(round(float(text)))
    except (ValueError, InvalidOperation):
        return 0


def is_skippable_name(name: str) -> bool:
    upper = name.strip().upper()
    if not upper:
        return True
    if upper in {"NAMES", "NAME", "TOTAL", "FEE BALANCES"}:
        return True
    if upper.startswith("EXCEL GRASSLAND"):
        return True
    if upper.startswith("TERM "):
        return True
    if re.match(r"^GR\s*\d", upper):
        return True
    if "SUBTOTAL" in upper or "TOTAL OUTSTANDING" in upper:
        return True
    return False


def find_name_col(ws, max_row: int = 12) -> tuple[int | None, int]:
    for row in range(1, max_row + 1):
        for col in range(1, ws.max_column + 1):
            value = str(ws.cell(row, col).value or "").strip().upper()
            if value in {"NAMES", "NAME"} or value.startswith("NAMES"):
                return row, col
    return None, 3


def find_balance_col(ws, header_row: int | None, start_row: int) -> int:
    if header_row:
        for col in range(1, ws.max_column + 1):
            value = str(ws.cell(header_row, col).value or "").strip().upper()
            if value == "BAL":
                return col

    for col in range(ws.max_column, 0, -1):
        hits = 0
        for row in range(start_row, min(start_row + 10, ws.max_row + 1)):
            cell = ws.cell(row, col).value
            if cell is None or cell == "":
                continue
            if isinstance(cell, (int, float)):
                hits += 1
                continue
            text = str(cell).strip().upper()
            if text in {"NIL", "NILL"} or re.fullmatch(r"-?\d+(\.\d+)?", text.replace(",", "")):
                hits += 1
        if hits >= 2:
            return col
    return ws.max_column


def parse_sheet(ws, sheet_name: str):
    header_row, name_col = find_name_col(ws)
    start_row = header_row + 1 if header_row else 2
    balance_col = find_balance_col(ws, header_row, start_row)
    rows = []

    for row in range(start_row, ws.max_row + 1):
        adm_raw = ws.cell(row, 1).value
        name_raw = ws.cell(row, name_col).value
        if not name_raw:
            continue
        name = re.sub(r"\s+", " ", str(name_raw)).strip()
        if is_skippable_name(name):
            continue
        if adm_raw is None or str(adm_raw).strip().upper() == "TOTAL":
            continue
        try:
            adm_no = str(int(float(adm_raw)))
        except (TypeError, ValueError):
            continue

        balance = parse_balance(ws.cell(row, balance_col).value)
        rows.append({
            "source_adm": adm_no,
            "name": name,
            "balance": balance,
            "sheet": sheet_name,
        })
    return rows


def assign_grassland_adm_nos(rows, existing_adms=None):
    """Append EG to every adm no; duplicate spreadsheet rows get EG2, EG3, etc."""
    existing_adms = existing_adms or set(
        Student.objects.filter(adm_no__endswith="EG").values_list("adm_no", flat=True)
    )
    assigned = set()
    seen = {}
    result = []

    for row in rows:
        src = row["source_adm"]
        seen[src] = seen.get(src, 0) + 1
        suffix = "" if seen[src] == 1 else str(seen[src])
        candidate = f"{src}EG{suffix}"

        while candidate in assigned or candidate in existing_adms:
            seen[src] += 1
            suffix = "" if seen[src] == 1 else str(seen[src])
            candidate = f"{src}EG{suffix}"

        row = {**row, "adm_no": candidate}
        assigned.add(candidate)
        result.append(row)

    extra_rows = sum(count - 1 for count in seen.values() if count > 1)
    return result, extra_rows


def resolve_class(school, sheet_name: str):
    mapping = SHEET_CLASS_MAP.get(sheet_name.lower())
    if not mapping:
        return None
    grade_name, stream_name = mapping
    grade = Grade.objects.filter(name=grade_name).first()
    if not grade:
        return None
    return Class.objects.filter(school=school, grade=grade, name=stream_name).first()


def clear_grassland_students(school):
    profiles = StudentProfile.objects.filter(school=school).select_related("student")
    student_ids = list(profiles.values_list("student_id", flat=True))
    profile_count = profiles.count()

    other = StudentProfile.objects.filter(student_id__in=student_ids).exclude(school=school).count()
    if other:
        raise RuntimeError(f"Abort: {other} Grassland students also linked to another school")

    with transaction.atomic():
        profiles.delete()
        Student.objects.filter(id__in=student_ids).delete()
    return profile_count


class Command(BaseCommand):
    help = "Import Excel Grassland students from Term 2 2026 workbook."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            default=str(DEFAULT_FILE),
            help="Path to the Term 2 2026 workbook (.xlsx)",
        )
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Delete existing Excel Grassland students before import",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and report only; do not write to the database",
        )

    def handle(self, *args, **options):
        path = Path(options["file"])
        if not path.exists():
            self.stderr.write(self.style.ERROR(f"File not found: {path}"))
            return

        school = School.objects.filter(name__iexact=SCHOOL_NAME).first()
        if not school:
            self.stderr.write(self.style.ERROR(f"{SCHOOL_NAME} not found. Run sync_excel_grassland_classes first."))
            return

        workbook = openpyxl.load_workbook(path, data_only=True)
        parsed_rows = []
        missing_classes = []

        for sheet_name in workbook.sheetnames:
            key = sheet_name.strip().lower()
            if key not in SHEET_CLASS_MAP:
                self.stdout.write(self.style.WARNING(f"Skipping unknown sheet: {sheet_name}"))
                continue

            klass = resolve_class(school, key)
            if not klass:
                missing_classes.append(sheet_name)
                continue

            rows = parse_sheet(workbook[sheet_name], sheet_name)
            for row in rows:
                row["class_obj"] = klass
                row["grade_name"] = klass.grade.name
            parsed_rows.extend(rows)
            self.stdout.write(f"{sheet_name}: {len(rows)} -> {klass.grade.name} / {klass.name}")

        if missing_classes:
            self.stderr.write(self.style.ERROR(f"Missing classes for sheets: {', '.join(missing_classes)}"))
            return

        self.stdout.write(f"Parsed {len(parsed_rows)} student rows")

        if options["dry_run"]:
            preview_rows, extra_rows = assign_grassland_adm_nos(parsed_rows, existing_adms=set())
            self.stdout.write(f"Would assign {len(preview_rows)} admission numbers (all end with EG)")
            if extra_rows:
                self.stdout.write(
                    self.style.WARNING(
                        f"{extra_rows} duplicate spreadsheet row(s) would get EG2/EG3 suffixes"
                    )
                )
            for row in preview_rows:
                if row["adm_no"] != f"{row['source_adm']}EG":
                    self.stdout.write(
                        f"  dup: {row['source_adm']} -> {row['adm_no']} ({row['name']}, {row['sheet']})"
                    )
            return

        if options["clear"]:
            cleared = clear_grassland_students(school)
            self.stdout.write(self.style.WARNING(f"Cleared {cleared} existing Grassland student profiles"))

        parsed_rows, extra_rows = assign_grassland_adm_nos(parsed_rows, existing_adms=set())
        if extra_rows:
            self.stdout.write(
                self.style.WARNING(
                    f"{extra_rows} duplicate spreadsheet row(s) assigned EG2/EG3 suffixes"
                )
            )
        self.stdout.write(f"Assigned {len(parsed_rows)} unique admission numbers (all end with EG)")

        created = 0
        with transaction.atomic():
            for row in parsed_rows:
                klass = row["class_obj"]
                first, middle, last = split_name(row["name"])
                gender = guess_gender(first)
                dob = DOB_BY_GRADE.get(klass.grade.name, date(2015, 6, 1))

                if Student.objects.filter(adm_no=row["adm_no"]).exists():
                    raise RuntimeError(f"Admission number already exists: {row['adm_no']}")

                student = Student.objects.create(
                    adm_no=row["adm_no"],
                    first_name=first,
                    middle_name=middle,
                    last_name=last,
                    date_of_birth=dob,
                    joined_date=JOINED,
                    gender=gender,
                    fee_category="day",
                    is_boarder=False,
                )

                StudentProfile.objects.create(
                    student=student,
                    school=school,
                    class_id=klass,
                    fee_balance=row["balance"],
                    status="Active",
                    discipline=100,
                )
                created += 1

        total = StudentProfile.objects.filter(school=school, status="Active").count()
        self.stdout.write(
            self.style.SUCCESS(
                f"Done. Created {created} students. Active Grassland students: {total}."
            )
        )
