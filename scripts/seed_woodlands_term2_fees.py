"""Parse d:\\FEE TERM 2 2026.xlsx and seed Excel Woodlands students.

Uses NAME + last BALANCE column only.
Places every grade into that grade's first stream (secretary will split later).
"""
import os
import re
import sys
from collections import OrderedDict
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "Excel.settings")

import django

django.setup()

from django.db import transaction
from django.db.models import Q

from core.models import School, Grade, Class, Student, StudentProfile

XLSX = Path(r"d:\FEE TERM 2 2026.xlsx")
SHEET = "TERM 2 2026"
DOB = date(2018, 1, 1)
JOINED = date(2026, 1, 5)

FEMALE_FIRST = {
    "TIFFANY", "KARRYN", "MAKYLA", "VALENTINE", "SHANAH", "FAITH", "FAUSTINA",
    "EPIPHANY", "ZAYLA", "MILA", "ADRIANA", "MARY", "MARYANN", "ANN", "ANNE",
    "ANNIE", "GRACE", "JOY", "JOYCE", "JANE", "JANET", "JULIET", "JULIA",
    "ESTHER", "ELIZABETH", "LIZ", "LIZA", "WANGARI", "WANJIRU", "WANJIKU",
    "NJERI", "NYAMBURA", "WAMBUI", "AWUOR", "ATENO", "AWINO", "CHEBET",
    "CHEMUTAI", "CHERONO", "JEPCHIRCHIR", "STACEY", "STACY", "ASHLEY",
    "REBECCA", "SARAH", "SARA", "HOPE", "YVONNE", "TASHA", "KIMBERLY",
    "ERICAH", "AUDREY", "SHARLYN", "NATASHA", "ANGELA", "IRENE", "CLARE",
    "CHLOE", "WHITNEY", "WINNIE", "ALEXIS", "SALMA", "LIBERTY", "SAMANTHA",
    "SASHA", "EVA", "TAMARA", "WANDA", "BUSHR", "BUSHRA", "PRISCAH", "PRISCA",
    "MICHELLE", "MITCHELLE", "LAURA", "SABRINA", "FAITH", "ABIGAEL", "HAJRA",
    "JOY", "TIFFANY", "DIANA", "TASHLEY", "SHIRLEEN", "JEAN", "WARDA", "AGNES",
    "PAMELA", "FAVOUR", "TRACY", "PRECIOUS", "PEACE", "IVY", "IVYNE", "CYNTHIA",
    "CATHERINE", "CATHY", "TEGAN", "TEGAN", "NOELLE", "NORAH", "NORA", "HANNAH",
    "HANNA", "LEILA", "LAYLA", "AALIYAH", "AALIYA", "IMANI", "ZARA", "ZARIA",
    "ZAWADI", "QUEEN", "PRINCESS", "BLESSING", "MERCY", "PATIENCE", "CHARITY",
    "HOPE", "PEACE", "JOY", "FAITH", "LOVE", "GIFT", "ANGEL", "ANGELINE",
    "ANGELICA", "BEATRICE", "BETTY", "ROSE", "ROSELYN", "ROSELINE", "LILIAN",
    "LILLIAN", "LUCY", "LUCIA", "SOPHIA", "SOFIA", "EMILY", "EMMA", "ELLA",
    "ELIANA", "ELIANA", "ARIA", "ARIANA", "MAYA", "MIA", "NOA", "NOAH",  # Noah often male - remove
}

# Remove ambiguous
FEMALE_FIRST.discard("NOAH")
FEMALE_FIRST.discard("NOA")


def normalize_grade(title):
    t = re.sub(r"\s+", " ", str(title or "").upper()).strip()
    if "PLAY GROUP" in t or "PLAYGROUP" in t:
        return "Play Group"
    if re.search(r"PP\s*1\b|PP1", t) and not re.search(r"PP\s*2\b|PP2", t):
        return "PP1"
    if re.search(r"\bPP\s*2\b|PP2", t):
        return "PP2"
    m = re.search(r"GRADE\s*(ONE|TWO|THREE|FOUR|FIVE|SIX|SEVEN|EIGHT|NINE|[1-9])", t)
    if not m:
        return None
    word = m.group(1)
    words = {
        "ONE": "1", "TWO": "2", "THREE": "3", "FOUR": "4", "FIVE": "5",
        "SIX": "6", "SEVEN": "7", "EIGHT": "8", "NINE": "9",
    }
    num = words.get(word, word)
    return f"Grade {num}"


def parse_balance(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return Decimal(str(int(round(float(v)))))
    s = str(v).replace(",", "").replace(" ", "").strip()
    if not s or s in ("-", "—", "N/A", "NA"):
        return Decimal("0")
    try:
        return Decimal(str(int(round(float(s)))))
    except Exception:
        return None


def split_name(full: str):
    parts = [p for p in full.strip().split() if p]
    if len(parts) == 1:
        return parts[0].title(), "", parts[0].title()
    if len(parts) == 2:
        return parts[0].title(), "", parts[1].title()
    return parts[0].title(), " ".join(parts[1:-1]).title(), parts[-1].title()


def first_stream_for(school, grade_name):
    """Prefer Longonot for multi-stream grades; else first by name/id."""
    grade = Grade.objects.filter(name=grade_name).first()
    if not grade:
        return None
    qs = Class.objects.filter(school=school, grade=grade)
    preferred = qs.filter(name__iexact="Longonot").first()
    if preferred:
        return preferred
    # Single-stream grades named after the grade
    same = qs.filter(name__iexact=grade_name).first()
    if same:
        return same
    return qs.order_by("id").first()


def parse_workbook(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Sheet {SHEET!r} not found. Have: {wb.sheetnames}")
    ws = wb[SHEET]

    current = None
    name_idx = bal_idx = None
    rows_by_grade = OrderedDict()

    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        b = vals[1] if len(vals) > 1 else None
        grade = normalize_grade(b)
        if grade and "NAME" not in str(b or "").upper():
            current = grade
            name_idx = bal_idx = None
            rows_by_grade.setdefault(current, OrderedDict())
            continue

        strs = [str(v).strip().upper() if v is not None else "" for v in vals]
        if "NAME" in strs:
            name_idx = strs.index("NAME")
            bal_idx = None
            for i, s in enumerate(strs):
                if s == "BALANCE":
                    bal_idx = i
            continue

        if current is None or name_idx is None or bal_idx is None:
            continue

        raw_name = vals[name_idx] if name_idx < len(vals) else None
        if not raw_name or not str(raw_name).strip():
            continue
        name = re.sub(r"\s+", " ", str(raw_name)).strip()
        if name.upper() in {"NAME", "TOTAL", "TOTALS", "GRAND TOTAL"}:
            continue
        bal = parse_balance(vals[bal_idx] if bal_idx < len(vals) else None)
        if bal is None:
            continue
        # Later duplicate sections overwrite earlier same-name rows in that grade
        rows_by_grade[current][name.casefold()] = (name, bal)

    return rows_by_grade


def run(dry_run=False):
    if not XLSX.exists():
        raise SystemExit(f"File not found: {XLSX}")

    school = School.objects.filter(name__iexact="Excel Woodlands").first()
    if not school:
        raise SystemExit("Excel Woodlands not found")

    parsed = parse_workbook(XLSX)
    print("Parsed from", XLSX.name, "sheet", SHEET)
    for g, od in parsed.items():
        print(f"  {g}: {len(od)} learners")

    created = updated = 0
    with transaction.atomic():
        for grade_name, od in parsed.items():
            klass = first_stream_for(school, grade_name)
            if not klass:
                raise SystemExit(f"No class/stream found for Woodlands {grade_name}")
            print(f"\n{grade_name} -> {klass.grade.name} / {klass.name}")
            for _key, (name, bal) in od.items():
                first, middle, last = split_name(name)
                gender = "female" if first.upper() in FEMALE_FIRST else "male"

                student = (
                    Student.objects.filter(
                        first_name__iexact=first,
                        last_name__iexact=last,
                        studentprofile__school=school,
                        studentprofile__class_id__grade__name=grade_name,
                    )
                    .filter(Q(adm_no__isnull=True) | Q(adm_no=""))
                    .first()
                )
                if not student:
                    # also match if already has adm later
                    student = Student.objects.filter(
                        first_name__iexact=first,
                        last_name__iexact=last,
                        studentprofile__school=school,
                        studentprofile__class_id__grade__name=grade_name,
                    ).first()

                if not student:
                    if dry_run:
                        print(f"  CREATE {name} bal={bal}")
                        created += 1
                        continue
                    student = Student(
                        adm_no=None,
                        first_name=first,
                        middle_name=middle,
                        last_name=last,
                        date_of_birth=DOB,
                        joined_date=JOINED,
                        gender=gender,
                        fee_category="day",
                        is_boarder=False,
                    )
                    student.save()
                    created += 1
                    tag = "Created"
                else:
                    if dry_run:
                        print(f"  UPDATE {name} bal={bal}")
                        updated += 1
                        continue
                    student.first_name = first
                    student.middle_name = middle
                    student.last_name = last
                    student.gender = gender
                    student.fee_category = "day"
                    student.is_boarder = False
                    student.save()
                    updated += 1
                    tag = "Updated"

                profile, _ = StudentProfile.objects.get_or_create(
                    student=student,
                    defaults={
                        "school": school,
                        "class_id": klass,
                        "fee_balance": bal,
                        "status": "Active",
                        "discipline": 100,
                    },
                )
                profile.school = school
                profile.class_id = klass
                profile.fee_balance = bal
                profile.status = "Active"
                profile.save()
                print(f"  {tag}: {student.get_full_name()} balance={profile.fee_balance}")

    print(f"\nDone. Created {created}, updated {updated}.")
    for grade_name in parsed:
        klass = first_stream_for(school, grade_name)
        n = StudentProfile.objects.filter(class_id=klass).count()
        print(f"  {klass.grade.name} / {klass.name}: {n}")


if __name__ == "__main__":
    dry = "--dry-run" in sys.argv
    run(dry_run=dry)
