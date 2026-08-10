"""Clear Excel Woodlands students only, then seed from bln.xlsx (final balances)."""
import os
import re
import sys
from collections import OrderedDict
from datetime import date
from decimal import Decimal
from pathlib import Path

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "Excel.settings")

import django

django.setup()

from django.db import transaction
from django.db.models import Q

from core.models import School, Grade, Class, Student, StudentProfile

XLSX = Path(os.environ.get("WOODLANDS_BLN_XLSX", r"c:\Users\USER\Downloads\bln.xlsx"))
SHEET = "Term 2 Balances"
DOB = date(2018, 1, 1)
JOINED = date(2026, 1, 5)

FEMALE_FIRST = {
    "TIFFANY", "KARRYN", "MAKYLA", "VALENTINE", "SHANAH", "FAITH", "FAUSTINA",
    "EPIPHANY", "ZAYLA", "MILA", "ADRIANA", "MARY", "MARYANN", "ANN", "ANNE",
    "ANNIE", "GRACE", "JOY", "JOYCE", "JANE", "JANET", "JULIET", "JULIA",
    "ESTHER", "ELIZABETH", "LIZ", "LIZA", "WANGARI", "WANJIRU", "WANJIKU",
    "NJERI", "NYAMBURA", "WAMBUI", "AWUOR", "ATENO", "AWINO", "CHEBET",
    "STACEY", "STACY", "ASHLEY", "REBECCA", "SARAH", "SARA", "HOPE", "YVONNE",
    "TASHA", "KIMBERLY", "AUDREY", "NATASHA", "ANGELA", "IRENE", "CLARE",
    "CHLOE", "WHITNEY", "WINNIE", "ALEXIS", "SALMA", "LIBERTY", "SAMANTHA",
    "SASHA", "EVA", "TAMARA", "WANDA", "BUSHRA", "PRISCAH", "PRISCA",
    "MICHELLE", "MITCHELLE", "LAURA", "SABRINA", "ABIGAEL", "ABIGAIL", "HAJRA",
    "DIANA", "TASHLEY", "SHIRLEEN", "JEAN", "WARDA", "AGNES", "PAMELA",
    "FAVOUR", "TRACY", "PRECIOUS", "PEACE", "IVY", "CYNTHIA", "CATHERINE",
    "HANNAH", "LEILA", "LAYLA", "IMANI", "ZARA", "ZARIA", "ZAWADI", "QUEEN",
    "PRINCESS", "BLESSING", "MERCY", "PATIENCE", "CHARITY", "ANGEL", "ANGELINE",
    "BEATRICE", "BETTY", "ROSE", "LILIAN", "LILLIAN", "LUCY", "SOPHIA", "SOFIA",
    "EMILY", "EMMA", "ELLA", "ELIANA", "ARIA", "ARIANA", "MAYA", "MIA",
    "CHARLOTTE", "NATALIE", "HELLEN", "BILHA", "PURITY", "SHANTEL", "VICTORIA",
    "MARGARET", "GERMAIN", "CHARITY", "GABRIELLA", "RUTH", "MILICENT", "PATRICIA",
    "ELIANA", "PHILOMENA", "LOISE", "QUEENSEY", "AVIVA", "MERCYANN", "SHARLENE",
    "NATALIA", "BELVER", "BLESSING", "TERRYANN", "BERNICE", "PHOEBE", "GENEVIEVE",
    "TASHLY", "OLIVE", "CAREN", "SHARON", "TREASURE", "TABITHA", "MONICAH", "LISA",
    "NAOMI", "SHIRLEY", "CHALLEN", "PRUDENCE", "MELISA", "PETRA", "JOANN", "WENDY",
    "VANESSA", "CHARLEEN", "WINFRED", "SWALHA", "JASMINE", "BLESSY", "STARLYN",
    "ALICE", "SHIPRA", "MAIDA", "GLORIA", "ELSIE", "BRIANNA", "ELLYANA", "MARYANN",
}


def normalize_grade(title):
    t = re.sub(r"\s+", " ", str(title or "").upper()).strip()
    if not t:
        return None
    if "PLAY GROUP" in t or "PLAYGROUP" in t:
        return "Play Group"
    if re.search(r"PP\s*1\b|^PP1$", t) and not re.search(r"PP\s*2\b|PP2", t):
        return "PP1"
    if re.search(r"PP\s*2\b|^PP2$", t):
        return "PP2"
    m = re.search(r"GRADE\s*(ONE|TWO|THREE|FOUR|FIVE|SIX|SEVEN|EIGHT|NINE|[1-9])", t)
    if not m:
        return None
    word = m.group(1)
    words = {
        "ONE": "1", "TWO": "2", "THREE": "3", "FOUR": "4", "FIVE": "5",
        "SIX": "6", "SEVEN": "7", "EIGHT": "8", "NINE": "9",
    }
    return f"Grade {words.get(word, word)}"


def parse_balance(v):
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, (int, float)):
        return Decimal(str(int(round(float(v)))))
    s = str(v).replace(",", "").replace(" ", "").strip()
    if not s or s in ("-", "—"):
        return Decimal("0")
    return Decimal(str(int(round(float(s)))))


def split_name(full: str):
    parts = [p for p in full.strip().split() if p]
    if len(parts) == 1:
        return parts[0].title(), "", parts[0].title()
    if len(parts) == 2:
        return parts[0].title(), "", parts[1].title()
    return parts[0].title(), " ".join(parts[1:-1]).title(), parts[-1].title()


def first_stream_for(school, grade_name):
    grade = Grade.objects.filter(name=grade_name).first()
    if not grade:
        return None
    qs = Class.objects.filter(school=school, grade=grade)
    preferred = qs.filter(name__iexact="Longonot").first()
    if preferred:
        return preferred
    same = qs.filter(name__iexact=grade_name).first()
    if same:
        return same
    return qs.order_by("id").first()


def parse_bln(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Sheet {SHEET!r} not found. Have: {wb.sheetnames}")
    ws = wb[SHEET]
    rows_by_grade = OrderedDict()
    current = None

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a and b is None:
            grade = normalize_grade(a)
            if grade:
                current = grade
                rows_by_grade.setdefault(current, OrderedDict())
                continue
        if not current or not a:
            continue
        name = re.sub(r"\s+", " ", str(a)).strip()
        if not name or name.upper() in {"STUDENT NAME", "NAME"}:
            continue
        if name.upper().startswith("EXCEL "):
            continue
        # Skip spreadsheet summary rows
        upper = name.upper()
        if "SUBTOTAL" in upper or "TOTAL OUTSTANDING" in upper or upper.startswith("SCHOOL-WIDE"):
            continue
        if re.search(r"\(\s*\d+\s+STUDENTS?\s*\)", upper):
            continue
        bal = parse_balance(b)
        rows_by_grade[current][name.casefold()] = (name, bal)
    return rows_by_grade


def clear_woodlands_only(school):
    """Delete learners belonging to Excel Woodlands only."""
    profiles = StudentProfile.objects.filter(school=school).select_related("student")
    student_ids = list(profiles.values_list("student_id", flat=True))
    profile_count = profiles.count()

    # Safety: never touch other schools' students
    other = StudentProfile.objects.filter(student_id__in=student_ids).exclude(school=school).count()
    if other:
        raise SystemExit(f"Abort: {other} selected students also linked to another school")

    deleted_students = 0
    with transaction.atomic():
        profiles.delete()
        # Delete students that no longer have any profile
        orphan_qs = Student.objects.filter(id__in=student_ids)
        deleted_students, _ = orphan_qs.delete()

    remaining = StudentProfile.objects.filter(school=school).count()
    academy = StudentProfile.objects.filter(school__name__iexact="Excel Academy").count()
    print(f"Cleared Woodlands profiles={profile_count}, student-delete-ops={deleted_students}")
    print(f"Woodlands remaining={remaining} | Academy still={academy}")
    if remaining:
        raise SystemExit("Woodlands clear incomplete")
    return profile_count


def run():
    if not XLSX.exists():
        raise SystemExit(f"File not found: {XLSX}")

    school = School.objects.filter(name__iexact="Excel Woodlands").first()
    if not school:
        raise SystemExit("Excel Woodlands not found")

    print("=== CLEAR Excel Woodlands ONLY ===")
    clear_woodlands_only(school)

    parsed = parse_bln(XLSX)
    print("\n=== SEED from", XLSX.name, "===")
    for g, od in parsed.items():
        print(f"  {g}: {len(od)}")

    created = 0
    with transaction.atomic():
        for grade_name, od in parsed.items():
            klass = first_stream_for(school, grade_name)
            if not klass:
                raise SystemExit(f"No class/stream for Woodlands {grade_name}")
            print(f"\n{grade_name} -> {klass.grade.name} / {klass.name}")
            for _key, (name, bal) in od.items():
                first, middle, last = split_name(name)
                gender = "female" if first.upper() in FEMALE_FIRST else "male"
                student = Student(
                    adm_no=None,
                    first_name=first,
                    middle_name=middle,
                    last_name=last,
                    date_of_birth=DOB,
                    joined_date=JOINED,
                    gender=gender,
                    fee_category="day",
                    is_boarder=False,
                )
                student.save()
                StudentProfile.objects.create(
                    student=student,
                    school=school,
                    class_id=klass,
                    fee_balance=bal,
                    status="Active",
                    discipline=100,
                )
                created += 1
                print(f"  Created: {student.get_full_name()} balance={bal}")

    print(f"\nDone. Created {created}.")
    for grade_name in parsed:
        klass = first_stream_for(school, grade_name)
        n = StudentProfile.objects.filter(class_id=klass).count()
        print(f"  {klass.grade.name} / {klass.name}: {n}")
    print("Academy still:", StudentProfile.objects.filter(school__name__iexact="Excel Academy").count())


if __name__ == "__main__":
    run()
